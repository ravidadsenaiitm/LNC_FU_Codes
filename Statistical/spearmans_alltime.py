import pandas as pd
import numpy as np
from scipy.stats import spearmanr
from statsmodels.stats.multitest import multipletests
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# === CONFIGURATION ===
input_file = "/Users/ravidadsena/Downloads/EFACTS_Crosssectional/Results/T1_analysis/Correlation_Analysis/Corrected/New_Correlation_28_06/specific_region_correlation/final_timeseries_correlation_imaging_T1_residual.xlsx"
output_file = "/Users/ravidadsena/Downloads/EFACTS_Crosssectional/Results/T1_analysis/Correlation_Analysis/Corrected/New_Correlation_28_06/specific_region_correlation/correlation_results_combined_all_sheets_structural_fdr.xlsx"

# Change this line as needed:
adjust_method = "fdr_bh"  # Options: "none", "fdr_bh", "bonferroni", "holm"

# === FORMATTING ===
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# === LOAD DATA ===
xls = pd.ExcelFile(input_file)
sheet_names = xls.sheet_names

# === CREATE OUTPUT WORKBOOK ===
wb = Workbook()
wb.remove(wb.active)

for sheet in sheet_names:
    print(f"\n=== Processing: {sheet} ===")
    df = xls.parse(sheet)

    if "handedness" not in df.columns:
        print(f"Skipping {sheet} (no 'handedness')")
        continue

    handedness_index = df.columns.get_loc("handedness")
    feature_cols = df.iloc[:, 4:handedness_index]
    variable_cols = df.iloc[:, handedness_index:]

    # Process handedness
    if "handedness" in variable_cols.columns:
        variable_cols["handedness"] = variable_cols["handedness"].apply(lambda x: 1 if x == "R" else 0)

    # Numeric conversion and NaN handling
    variable_cols = variable_cols.apply(pd.to_numeric, errors='coerce')
    variable_cols = variable_cols.fillna(variable_cols.mean())

    n_features, n_vars = feature_cols.shape[1], variable_cols.shape[1]
    corr_matrix = np.full((n_features, n_vars), np.nan)
    pval_matrix = np.full((n_features, n_vars), np.nan)

    # === COMPUTE CORRELATIONS ===
    for i in range(n_features):
        for j in range(n_vars):
            x, y = feature_cols.iloc[:, i], variable_cols.iloc[:, j]
            valid = (~x.isna()) & (~y.isna())
            if valid.sum() >= 3:
                corr, pval = spearmanr(x[valid], y[valid])
                corr_matrix[i, j], pval_matrix[i, j] = corr, pval

    # ✅ Confirm: Correlation values (before p-adjustment)
    print("✅ Correlation Check (before p-adjust):", np.nanmean(corr_matrix).round(3))

    # === ADJUST P-VALUES ONLY ===
    if adjust_method.lower() != "none":
        flat_pvals = pval_matrix.flatten()
        mask = ~np.isnan(flat_pvals)
        adjusted = np.full_like(flat_pvals, np.nan)
        adjusted[mask] = multipletests(flat_pvals[mask], method=adjust_method.lower())[1]
        pval_matrix = adjusted.reshape(pval_matrix.shape)

    # ✅ Confirm: Correlation values unchanged after p-adjust
    print("✅ Correlation Check (after p-adjust):", np.nanmean(corr_matrix).round(3))

    # === FORMAT OUTPUT ===
    corr_df = pd.DataFrame(corr_matrix, index=feature_cols.columns, columns=variable_cols.columns)
    pval_df = pd.DataFrame(pval_matrix, index=feature_cols.columns, columns=variable_cols.columns)

    formatted_corr = corr_df.copy().astype(str)
    for i in range(n_features):
        for j in range(n_vars):
            val, pval = corr_matrix[i, j], pval_matrix[i, j]
            if np.isnan(val):
                formatted_corr.iloc[i, j] = "NA"
            elif pval < 0.05:
                formatted_corr.iloc[i, j] = f"**{val:.2f}"
            else:
                formatted_corr.iloc[i, j] = f"{val:.2f}"

    # === WRITE TO EXCEL SHEET ===
    ws = wb.create_sheet(title=sheet)
    ws.append(["Feature"] + list(formatted_corr.columns))
    for idx, row in formatted_corr.iterrows():
        ws.append([idx] + list(row))

    # HIGHLIGHT SIGNIFICANT CORRELATIONS
    for i in range(2, 2 + formatted_corr.shape[0]):
        for j in range(2, 2 + formatted_corr.shape[1]):
            if "**" in str(ws.cell(row=i, column=j).value):
                ws.cell(row=i, column=j).fill = yellow_fill

    # === ADD P-VALUE MATRIX SHEET ===
    ws_p = wb.create_sheet(title=f"{sheet}_p")
    ws_p.append(["Feature"] + list(pval_df.columns))
    for idx, row in pval_df.iterrows():
        ws_p.append([idx] + list(np.round(row.values, 4)))

    for i in range(2, 2 + pval_df.shape[0]):
        for j in range(2, 2 + pval_df.shape[1]):
            try:
                if float(ws_p.cell(row=i, column=j).value) < 0.05:
                    ws_p.cell(row=i, column=j).fill = yellow_fill
            except:
                continue

# === SAVE OUTPUT ===
wb.save(output_file)
print(f"\n✅ All results saved to: {output_file}")
